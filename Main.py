
from tkinter import *
from tkinter import filedialog



import zipfile
import os
import shutil

from openpyxl import load_workbook,Workbook

import openpyxl
from OpenPyXLHelperFunctions import data_from_range, data_to_range

import pandas as pd





#############################Functions

def do():
	print(mathFile.get())
	print(fileDir)
	CopyFiles(mathFile.get(), fileDir)

def Exit():
	root.destory()

def get_file_dir():
	global fileDir
	fileDir = filedialog.askdirectory()	

def get_developer_file_list(directory,math,developerExtension):
    os.chdir(directory)
    filez = os.listdir()    
    files = []
    
    for file in filez:        
        try:
            if file.split('.')[1] in developerExtension and file[0] != '~'  and file != str(math):
                files.append(file)
        except:
            pass
            
    return(files)

def CopyFiles(mathFile, fileDir):

    math = mathFile

    developerExtension = ['xlsx','xlsm']

    ranges = pd.read_table("Ranges.txt")
    numRanges = ranges.shape[0]
    
    #gets the names of all of the developer files. 
    files = get_developer_file_list(fileDir,math,developerExtension)    
    print(files)
    
    #loops through all of the developer files, renames our math file, 
    #copies and pastes each range that
    #we need and saves the workbook
    for file in files:        
        srcBookName = file
        tgtBookName = 'DGE - ' + srcBookName
        
        shutil.copy(math,tgtBookName)
        
        srcBook = load_workbook(srcBookName, data_only = True)
        tgtBook = load_workbook(tgtBookName)
    
        for col in range(numRanges):    
                srcRngString = ranges['From'][col]                
                tgtRngString = ranges['To'][col]
                print(srcRngString, ',', tgtRngString)
                srcData = data_from_range(srcRngString, srcBook)
                data_to_range(srcData,tgtRngString,tgtBook)  
                
        srcBook.save(file)
        
        #selects the first sheet(hopefully the summary sheet) before saving
        tgtBook.active = 0
        tgtBook.save('DGE - ' + file)
    print('You are all done copying')


##############################

root = Tk()
root.geometry("400x400+0+0")
root.title('Main Window')


mathFile = StringVar()
fileDir = StringVar()




f1 = Frame(root)
f1.pack()

#####get the name of the math file
mathFileLabel = Label(f1, text = 'Please enter the name of the math file')
mathFileLabel.grid(row =0, column = 0)
mathFileEntry = Entry(f1, textvariable = mathFile)
mathFileEntry.grid(row = 0, column = 1)


############ get the file directory

fileDirButton = Button(f1, command = lambda: get_file_dir(), text = "Let's figure out where the files are located")
fileDirButton.grid(row = 1, column = 0,columnspan = 2, sticky=E+W)


###########################################


finalButton = Button(f1, text = 'Press Me When you are done', command = lambda: do())
finalButton.grid(row = 10,column = 0)




root.mainloop()